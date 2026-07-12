from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from models.vehicle import Vehicle


def export_vehicles_excel(db: Session) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicles"

    headers = [
        "Registration Number",
        "Model",
        "Type",
        "Capacity (kg)",
        "Odometer (km)",
        "Current location",
        "Status",
    ]

    ws.append(headers)

    # Make header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    vehicles = db.scalars(select(Vehicle).order_by(Vehicle.registration_number)).all()

    for vehicle in vehicles:
        ws.append(
            [
                vehicle.registration_number,
                vehicle.model_name,
                vehicle.type.value,
                vehicle.max_load_capacity,
                vehicle.odometer,
                vehicle.current_location,
                vehicle.status.value,
            ]
        )

    # Auto-size columns
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)

        ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 3

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    return buffer
